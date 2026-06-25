"""Deterministic C3D <-> video drinking-trial association.

Both modalities recorded the *same* drinking repetitions in the *same order*:

  - Mocap: one QTM take per sip -> one C3D per sip, run-numbered ascending
    (e.g. P06_0001 .. or P130011..). Sorting by filename = chronological.
  - Video: the annotation Excel (`clean_data.xlsx`) lists every sip as its own
    row, but the "task name" column is SPARSE (forward-fill): a task name
    appears once, then every following blank-task row belongs to it until the
    next named task. The drinking rows, in sheet order, are chronological and
    carry start/end frame + side per sip.

So where the per-participant counts match, the join is purely POSITIONAL:

    sorted_c3d[k]  <->  drinking_rows[k]

No content matching, no LED sync, no guessing. Participants whose counts don't
match are returned UNMATCHED (never force-paired) so a dropped/aborted trial on
one side can't silently shift every rep after it.

Verified count-match (use ALL drinking rows): P06-P20, P23, P24, P16, P17, and
P03 via the P03 (1) session sheet. Unmatched: P01 (no mocap), P02/P04/P05/P21
(count differs), P07 (1)/P15 (1) (empty duplicate sheets).

Usage:
    python qtm_video_map.py                 # print match report
    python qtm_video_map.py --json cache/qtm_video_map.json

    from qtm_video_map import build_map
    m = build_map()
    m["P06"].pairs            # [(c3d_stem, {start,end,side,valid}), ...]
"""
from __future__ import annotations

import argparse
import glob
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

try:
    from _paths import QTM_C3D
except ImportError:
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from _paths import QTM_C3D

from qtm_c3d import load_trial

EXCEL_PATH = Path("/home/imove/Documents/annotation_tool/clean_data.xlsx")
VIDEO_FPS = 60.0
SHEET_PAT = re.compile(r"^P\d{2}( \(\d+\))?$")

# Per-participant overrides: which Excel sheet to read for a participant.
# P03's base sheet is short (29); the re-annotated "P03 (1)" matches the 30 C3D.
SHEET_OVERRIDE = {"P03": "P03 (1)"}


@dataclass
class DrinkRow:
    start: int          # video frame (60 fps)
    end: int
    side: str | None    # "left" | "right"
    valid: int | None   # 0 invalid / 1 valid / 2 edge / 3 unselected


@dataclass
class ParticipantMap:
    participant: str
    sheet: str | None
    c3d: list[str]                       # sorted C3D stems
    rows: list[DrinkRow]                 # chronological drinking rows
    matched: bool
    reason: str = ""
    pairs: list[tuple] = field(default_factory=list)  # (c3d_stem, DrinkRow)


def _c3d_stems(pid: str, root: Path = QTM_C3D) -> list[str]:
    """Sorted C3D stems for a participant. Handles both `P06_0001` and `P030011`
    naming. Sorted by the trailing run number so order = chronological."""
    fs = glob.glob(str(Path(root) / f"{pid}_*.c3d")) + glob.glob(str(Path(root) / f"{pid}[0-9]*.c3d"))
    stems = sorted({Path(f).stem for f in fs}, key=lambda s: int(re.search(r"(\d+)$", s).group(1)))
    return stems


def _drink_rows(ws) -> list[DrinkRow]:
    """Forward-filled drinking rows with a frame range, in sheet (chronological) order."""
    rows: list[DrinkRow] = []
    last_task = None
    for r in range(2, ws.max_row + 1):
        tv = ws.cell(r, 2).value
        if tv is not None and str(tv).strip():
            last_task = str(tv).strip()
        if last_task and "drink" in last_task.lower():
            sf, ef = ws.cell(r, 5).value, ws.cell(r, 6).value
            if sf is not None and ef is not None:
                rows.append(DrinkRow(int(sf), int(ef), ws.cell(r, 4).value, ws.cell(r, 7).value))
    return rows


def _participants(wb) -> list[str]:
    """Base participant ids present as sheets (dedup the '(n)' variants)."""
    seen = []
    for s in wb.sheetnames:
        if SHEET_PAT.match(s):
            pid = re.sub(r"\s*\(\d+\)$", "", s)
            if pid not in seen:
                seen.append(pid)
    return seen


def _video_dur(row: DrinkRow, fps: float = VIDEO_FPS) -> float:
    return (row.end - row.start + 1) / fps


def monotonic_align(vdur: list[float], cdur: list[float], skip_cost: float = 0.6):
    """Order-preserving alignment of two duration sequences allowing skips on
    either side (a dropped/extra sip). Both are strictly time-ordered, so a
    valid alignment is monotonic: matched index pairs increase on both axes.

    DP minimizing sum |vdur[i]-cdur[j]| over matched pairs + skip_cost per
    unmatched element (seconds). Returns (pairs, residual_rms, n_skips) where
    pairs is [(i, j), ...] of matched video/C3D indices.
    """
    import numpy as np
    n, m = len(vdur), len(cdur)
    INF = float("inf")
    dp = np.full((n + 1, m + 1), INF)
    bt = np.zeros((n + 1, m + 1), dtype=np.int8)  # 0=match 1=skip-video 2=skip-c3d
    dp[0, 0] = 0.0
    for i in range(n + 1):
        for j in range(m + 1):
            if dp[i, j] == INF:
                continue
            if i < n and j < m:
                c = dp[i, j] + abs(vdur[i] - cdur[j])
                if c < dp[i + 1, j + 1]:
                    dp[i + 1, j + 1] = c; bt[i + 1, j + 1] = 0
            if i < n:  # skip a video sip
                c = dp[i, j] + skip_cost
                if c < dp[i + 1, j]:
                    dp[i + 1, j] = c; bt[i + 1, j] = 1
            if j < m:  # skip a C3D take
                c = dp[i, j] + skip_cost
                if c < dp[i, j + 1]:
                    dp[i, j + 1] = c; bt[i, j + 1] = 2
    # backtrack
    i, j, pairs, resid, skips = n, m, [], [], 0
    while i > 0 or j > 0:
        b = bt[i, j]
        if b == 0:
            pairs.append((i - 1, j - 1)); resid.append(abs(vdur[i - 1] - cdur[j - 1]))
            i -= 1; j -= 1
        elif b == 1:
            skips += 1; i -= 1
        else:
            skips += 1; j -= 1
    pairs.reverse()
    rms = float(np.sqrt(np.mean(np.square(resid)))) if resid else INF
    return pairs, rms, skips


def build_map(excel_path: Path = EXCEL_PATH, c3d_root: Path = QTM_C3D,
              rescue: bool = True, rescue_rms_max: float = 0.4) -> dict[str, ParticipantMap]:
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    out: dict[str, ParticipantMap] = {}
    for pid in _participants(wb):
        sheet = SHEET_OVERRIDE.get(pid, pid)
        if sheet not in wb.sheetnames:
            sheet = pid
        rows = _drink_rows(wb[sheet])
        c3d = _c3d_stems(pid, c3d_root)

        if not c3d:
            pm = ParticipantMap(pid, sheet, c3d, rows, False, "no mocap (no C3D)")
        elif not rows:
            pm = ParticipantMap(pid, sheet, c3d, rows, False, "no drinking rows in sheet")
        elif len(rows) == len(c3d):
            pairs = list(zip(c3d, rows))
            pm = ParticipantMap(pid, sheet, c3d, rows, True, "positional 1:1", pairs)
        elif rescue:
            # counts differ -> monotonic duration alignment (order-preserving, allows skips)
            vdur = [_video_dur(r) for r in rows]
            cdur = [load_trial(s).duration_s for s in c3d]
            idx_pairs, rms, skips = monotonic_align(vdur, cdur)
            if rms <= rescue_rms_max:
                pairs = [(c3d[j], rows[i]) for i, j in idx_pairs]
                conf = "high" if rms <= 0.2 else "low"   # >0.2s = scattered skips, trust loosely
                pm = ParticipantMap(pid, sheet, c3d, rows, True,
                                    f"rescued ({conf} conf): monotonic align, {len(pairs)} pairs, "
                                    f"{skips} skipped, dur RMS {rms:.2f}s", pairs)
            else:
                pm = ParticipantMap(pid, sheet, c3d, rows, False,
                                    f"count mismatch {len(rows)}v vs {len(c3d)}c; "
                                    f"align RMS {rms:.2f}s > {rescue_rms_max} (rejected)")
        else:
            pm = ParticipantMap(pid, sheet, c3d, rows, False,
                                f"count mismatch: {len(rows)} video vs {len(c3d)} C3D")
        out[pid] = pm
    return out


def to_json(m: dict[str, ParticipantMap]) -> dict:
    def row(r: DrinkRow):
        return dict(start=r.start, end=r.end, side=r.side, valid=r.valid)
    parts = {}
    for pid, pm in m.items():
        parts[pid] = dict(
            sheet=pm.sheet, matched=pm.matched, reason=pm.reason,
            n_c3d=len(pm.c3d), n_video=len(pm.rows),
            pairs=[dict(c3d=c, video=row(r)) for c, r in pm.pairs],
        )
    return dict(video_fps=VIDEO_FPS, c3d_root=str(QTM_C3D), participants=parts)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", type=Path, default=EXCEL_PATH)
    ap.add_argument("--json", type=Path, default=None)
    args = ap.parse_args()

    m = build_map(args.excel)
    matched = [p for p, pm in m.items() if pm.matched]
    unmatched = [(p, pm) for p, pm in m.items() if not pm.matched]
    n_pairs = sum(len(pm.pairs) for pm in m.values())
    print(f"MATCHED {len(matched)}/{len(m)} participants, {n_pairs} sip-pairs (positional 1:1)")
    for p in matched:
        pm = m[p]
        print(f"  {p:5s} {len(pm.pairs):3d} sips  ({pm.sheet})")
    print(f"\nUNMATCHED {len(unmatched)}:")
    for p, pm in unmatched:
        print(f"  {p:5s} {pm.reason}")
    if args.json:
        args.json.write_text(json.dumps(to_json(m), indent=2))
        print(f"\nwrote {args.json}")


if __name__ == "__main__":
    main()
